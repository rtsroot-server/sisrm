import streamlit as st
import pandas as pd
import io
import re
import os
import unicodedata
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 1. Configuração da página - UI/UX Premium (Modo Wide)
st.set_page_config(
    page_title="Radar Político", 
    page_icon="🗺️", 
    layout="wide",
    initial_sidebar_state="auto"
)

# 2. CSS Customizado - Inspirado no Design "Gestão Online / AdminLTE"
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Cor de fundo do sistema (Cinza Claro) */
    .stApp { background-color: #ecf0f5; }
    h1, h2, h3 { color: #333333; font-family: 'Segoe UI', sans-serif; text-align: center; }
    
    /* Botões Padrão Premium */
    .stButton>button { 
        width: 100%; border-radius: 4px; font-weight: bold; 
        background-color: #3c8dbc; color: white; border: none;
        padding: 10px 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); transition: all 0.2s ease;
    }
    .stButton>button:hover {
        background-color: #367fa9; transform: translateY(-2px); box-shadow: 0 4px 6px rgba(0,0,0,0.15);
    }
    
    /* Caixa de Upload */
    [data-testid="stFileUploadDropzone"] {
        border-radius: 8px; border: 2px dashed #3c8dbc; background-color: #ffffff;
        padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    /* --- SIDEBAR CUSTOMIZADA (Dark Blue) --- */
    [data-testid="stSidebar"] {
        background-color: #222d32 !important;
    }
    [data-testid="stSidebar"] * {
        color: #b8c7ce !important;
    }
    [data-testid="stSidebar"] h2 {
        color: #ffffff !important;
        font-weight: bold;
    }
    /* Estilizando o menu de navegação na sidebar */
    div[role="radiogroup"] > label {
        padding: 10px;
        border-radius: 4px;
        transition: background-color 0.3s;
    }
    div[role="radiogroup"] > label:hover {
        background-color: #1e282c;
        color: #ffffff !important;
    }
    
    /* --- CARDS DO DASHBOARD (Estilo AdminLTE) --- */
    .small-box {
        border-radius: 4px;
        position: relative;
        display: block;
        margin-bottom: 20px;
        box-shadow: 0 1px 1px rgba(0,0,0,0.1);
        color: #fff;
        text-align: left;
    }
    .small-box > .inner {
        padding: 20px;
    }
    .small-box h3 {
        font-size: 45px;
        font-weight: bold;
        margin: 0 0 10px 0;
        white-space: nowrap;
        padding: 0;
        color: #fff;
        text-align: left;
    }
    .small-box p {
        font-size: 18px;
        margin-bottom: 0;
        color: #fff;
    }
    .small-box-footer {
        position: relative;
        text-align: center;
        padding: 5px 0;
        color: rgba(255,255,255,0.8);
        display: block;
        z-index: 10;
        background: rgba(0,0,0,0.15);
        text-decoration: none;
        font-size: 14px;
        border-bottom-left-radius: 4px;
        border-bottom-right-radius: 4px;
    }
    /* Cores fiéis à imagem */
    .bg-aqua { background-color: #00c0ef !important; }
    .bg-yellow { background-color: #f39c12 !important; }
    .bg-green { background-color: #00a65a !important; }
    .bg-red { background-color: #dd4b39 !important; }
    </style>
""", unsafe_allow_html=True)

# --- NOMES DOS ARQUIVOS DE BANCO DE DADOS (MEMÓRIA PERMANENTE) ---
MASTER_DB_FILE = "banco_dados_radar.csv"
CASTRACAO_DB_FILE = "banco_castracao_radar.csv"

# --- FUNÇÕES DE LÓGICA DE DADOS ---

def formatar_telefone(numero):
    if pd.isna(numero) or str(numero).strip() == "":
        return ""
    numeros_limpos = re.sub(r'\D', '', str(numero))
    if len(numeros_limpos) == 0:
        return ""
    if len(numeros_limpos) == 11:
        return f"({numeros_limpos[:2]}) {numeros_limpos[2:7]}-{numeros_limpos[7:]}"
    elif len(numeros_limpos) == 10:
        return f"({numeros_limpos[:2]}) 9{numeros_limpos[2:6]}-{numeros_limpos[6:]}"
    elif len(numeros_limpos) == 9:
        return f"(21) {numeros_limpos[:5]}-{numeros_limpos[5:]}"
    elif len(numeros_limpos) == 8:
        return f"(21) 9{numeros_limpos[:4]}-{numeros_limpos[4:]}"
    else:
        return numeros_limpos

def limpar_nome_e_bairro(nome_raw, bairro_raw):
    nome = str(nome_raw).strip()
    bairro = str(bairro_raw).strip()
    
    if " - " in nome or "-" in nome:
        partes = nome.split("-", 1)
        nome = partes[0].strip()
        bairro_extraido = partes[1].strip()
        if not bairro or bairro.lower() == 'nan':
            bairro = bairro_extraido

    bairro_lower = bairro.lower()
    bairro_invalido = False
    bairro_sem_espacos = bairro.replace(" ", "")
    
    if not bairro or bairro_lower == 'nan':
        bairro_invalido = True
    elif re.search(r'\d', bairro): 
        bairro_invalido = True
    elif len(bairro_sem_espacos) <= 3: 
        bairro_invalido = True
    elif "." in bairro and len(bairro_sem_espacos) <= 5: 
        bairro_invalido = True
    else:
        palavras_proibidas = ['gato', 'cachorro', 'cadela', 'adotou', 'adotante', 'devolveu', 'macho', 'fêmea', 'femea', 'filhote', 'pet', 'aa.c']
        for palavra in palavras_proibidas:
            if palavra in bairro_lower:
                bairro_invalido = True
                break
                
    if bairro_invalido:
        bairro = "Sem Bairro"
        
    bairro_norm = ''.join(c for c in unicodedata.normalize('NFD', bairro) if unicodedata.category(c) != 'Mn')
    bairro_norm = bairro_norm.replace('.', '. ')
    bairro_norm = re.sub(r'\s+', ' ', bairro_norm).strip().title()
        
    return nome.title(), bairro_norm

def processar_planilha(df):
    tel_col = None
    nome_col = None
    bairro_col = None
    agendamento_col = None
    
    for c in df.columns:
        c_lower = str(c).lower().strip()
        if c_lower in ['telefone', 'celular', 'contato', 'numero']:
            tel_col = c
        elif 'nome' in c_lower or 'cliente' in c_lower:
            nome_col = c
        elif c_lower in ['região', 'regiao', 'bairro', 'unnamed: 1']:
            bairro_col = c
        elif 'agendamento' in c_lower or 'visita' in c_lower:
            agendamento_col = c

    dados_padronizados = []
    
    for _, row in df.iterrows():
        n_raw = str(row[nome_col]) if nome_col and not pd.isna(row[nome_col]) else "Sem Nome"
        b_raw = str(row[bairro_col]) if bairro_col and not pd.isna(row[bairro_col]) else ""
        t_raw = str(row[tel_col]) if tel_col and not pd.isna(row[tel_col]) else ""
        a_raw = str(row[agendamento_col]) if agendamento_col and not pd.isna(row[agendamento_col]) else ""
        
        if a_raw.lower() == 'nan':
            a_raw = ""
            
        nome_limpo, bairro_limpo = limpar_nome_e_bairro(n_raw, b_raw)
        tel_formatado = formatar_telefone(t_raw)
        
        if tel_formatado != "":
            dados_padronizados.append({
                "Região/Bairro": bairro_limpo,
                "Nome": nome_limpo,
                "Telefone": tel_formatado,
                "Agendamento de Visita": a_raw
            })
            
    return pd.DataFrame(dados_padronizados)

def gerar_excel_final(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        bairros = df['Região/Bairro'].unique()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3c8dbc", end_color="3c8dbc", fill_type="solid")
        abas_usadas = {} 
        
        for bairro in bairros:
            nome_limpo_excel = re.sub(r'[\\/*?:\[\]]', '-', str(bairro)).strip()
            nome_aba_base = nome_limpo_excel[:31] if nome_limpo_excel[:31] else "Sem Bairro"
                
            nome_aba = nome_aba_base
            contador = 1
            while nome_aba in abas_usadas.values():
                sufixo = str(contador)
                nome_aba = nome_aba_base[:31 - len(sufixo)] + sufixo
                contador += 1
                
            abas_usadas[bairro] = nome_aba
                
            df_bairro = df[df['Região/Bairro'] == bairro].copy()
            if 'Agendamento de Visita' not in df_bairro.columns:
                df_bairro['Agendamento de Visita'] = "" 
            df_bairro = df_bairro[['Região/Bairro', 'Nome', 'Telefone', 'Agendamento de Visita']]
            df_bairro.to_excel(writer, index=False, sheet_name=nome_aba)
            
            worksheet = writer.sheets[nome_aba]
            for col_num, column_title in enumerate(df_bairro.columns, 1):
                col_letra = get_column_letter(col_num)
                if column_title == 'Nome': worksheet.column_dimensions[col_letra].width = 35
                elif column_title == 'Agendamento de Visita': worksheet.column_dimensions[col_letra].width = 40
                else: worksheet.column_dimensions[col_letra].width = 25
                
                celula = worksheet.cell(row=1, column=col_num)
                celula.font = header_font
                celula.fill = header_fill
                celula.alignment = Alignment(horizontal='center', vertical='center')

    return output.getvalue()

def gerar_excel_unico(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3c8dbc", end_color="3c8dbc", fill_type="solid")
        
        df_completo = df.copy()
        if 'Agendamento de Visita' not in df_completo.columns:
            df_completo['Agendamento de Visita'] = "" 
            
        df_completo.insert(0, 'Nº', range(1, len(df_completo) + 1))
        
        df_completo = df_completo[['Nº', 'Região/Bairro', 'Nome', 'Telefone', 'Agendamento de Visita']]
        df_completo.to_excel(writer, index=False, sheet_name="Dados Completos")
        
        worksheet = writer.sheets["Dados Completos"]
        for col_num, column_title in enumerate(df_completo.columns, 1):
            col_letra = get_column_letter(col_num)
            
            if column_title == 'Nº': worksheet.column_dimensions[col_letra].width = 8
            elif column_title == 'Nome': worksheet.column_dimensions[col_letra].width = 35
            elif column_title == 'Agendamento de Visita': worksheet.column_dimensions[col_letra].width = 40
            else: worksheet.column_dimensions[col_letra].width = 25
            
            celula = worksheet.cell(row=1, column=col_num)
            celula.font = header_font
            celula.fill = header_fill
            celula.alignment = Alignment(horizontal='center', vertical='center')

    return output.getvalue()

# --- GERENCIAMENTO DE ESTADO GERAL E PERSISTÊNCIA ---

if 'df_final' not in st.session_state:
    if os.path.exists(MASTER_DB_FILE):
        st.session_state.df_final = pd.read_csv(MASTER_DB_FILE, dtype=str).fillna("")
    else:
        st.session_state.df_final = None

if 'df_castracao' not in st.session_state:
    if os.path.exists(CASTRACAO_DB_FILE):
        try:
            df_c = pd.read_csv(CASTRACAO_DB_FILE, dtype=str).fillna("")
            if "Tipo" not in df_c.columns:
                st.session_state.df_castracao = pd.DataFrame(columns=["Tipo", "Nome", "Telefone"])
            else:
                st.session_state.df_castracao = df_c
        except:
            st.session_state.df_castracao = pd.DataFrame(columns=["Tipo", "Nome", "Telefone"])
    else:
        st.session_state.df_castracao = pd.DataFrame(columns=["Tipo", "Nome", "Telefone"])

if 'uploader_key' not in st.session_state:
    st.session_state.uploader_key = 0


# --- NAVEGAÇÃO LATERAL ---
with st.sidebar:
    col_img1, col_img2, col_img3 = st.columns([1, 2, 1])
    with col_img2:
        try:
            st.image("logo.png", use_container_width=True)
        except:
            pass
            
    st.markdown("<h2 style='text-align: center; margin-top:-10px;'>RADAR POLÍTICO</h2>", unsafe_allow_html=True)
    st.markdown("---")
    
    menu_selecionado = st.radio(
        "Navegação do Sistema:",
        ["📈 Dashboard", "🔄 Processamento de Dados", "🐾 Controle de Castração"]
    )

# ==========================================
# MÓDULO 1: DASHBOARD
# ==========================================
if menu_selecionado == "📈 Dashboard":
    st.title("Painel Gerencial")
    st.markdown("---")
    
    if st.session_state.df_final is not None and not st.session_state.df_final.empty:
        df = st.session_state.df_final
        
        if 'Agendamento de Visita' not in df.columns:
            df['Agendamento de Visita'] = ""
            
        total_cadastros = len(df)
        total_agendamentos = len(df[df['Agendamento de Visita'].str.strip() != ""])
        total_bairros = len(df['Região/Bairro'].unique())
        
        # Filtro de Bairro Superior
        lista_bairros = sorted(df['Região/Bairro'].unique())
        bairro_selecionado = st.selectbox("📌 Filtrar dados por Bairro:", lista_bairros)
        total_no_bairro = len(df[df['Região/Bairro'] == bairro_selecionado])
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Layout de 4 Colunas para os Cards
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f'''
            <div class="small-box bg-aqua">
                <div class="inner">
                    <h3>{total_cadastros}</h3>
                    <p>Total de Contatos</p>
                </div>
                <div class="small-box-footer">Base Geral Atualizada</div>
            </div>
            ''', unsafe_allow_html=True)
            
        with col2:
            st.markdown(f'''
            <div class="small-box bg-yellow">
                <div class="inner">
                    <h3>{total_agendamentos}</h3>
                    <p>Agendamentos</p>
                </div>
                <div class="small-box-footer">Visitas Marcadas</div>
            </div>
            ''', unsafe_allow_html=True)
            
        with col3:
            st.markdown(f'''
            <div class="small-box bg-green">
                <div class="inner">
                    <h3>{total_bairros}</h3>
                    <p>Regiões Atendidas</p>
                </div>
                <div class="small-box-footer">Mapeamento Geográfico</div>
            </div>
            ''', unsafe_allow_html=True)
            
        with col4:
            st.markdown(f'''
            <div class="small-box bg-red">
                <div class="inner">
                    <h3>{total_no_bairro}</h3>
                    <p>{bairro_selecionado}</p>
                </div>
                <div class="small-box-footer">Contatos neste Bairro</div>
            </div>
            ''', unsafe_allow_html=True)
            
    else:
        st.info("Sua base de dados está vazia no momento. Acesse o menu 'Processamento de Dados' para anexar suas planilhas.")

# ==========================================
# MÓDULO 2: PROCESSAMENTO DE DADOS
# ==========================================
elif menu_selecionado == "🔄 Processamento de Dados":
    st.title("Processamento de Dados")
    st.write("Unifique novas planilhas com a sua base salva automaticamente.")

    tamanho_atual = len(st.session_state.df_final) if st.session_state.df_final is not None else 0
    st.info(f"💾 **Status:** O seu banco de dados atual possui {tamanho_atual} contatos salvos.")

    col1, col2 = st.columns([3, 1])

    with col1:
        uploaded_files = st.file_uploader(
            "Toque ou arraste as NOVAS planilhas aqui", 
            type=["xlsx", "csv", "txt"], 
            accept_multiple_files=True,
            key=str(st.session_state.uploader_key)
        )

    with col2:
        st.markdown("<div style='margin-top: 35px;'></div>", unsafe_allow_html=True)
        if st.button("🗑️ Limpar Arquivos"):
            st.session_state.uploader_key += 1 
            st.rerun()

    if uploaded_files:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚀 Processar e Salvar no Banco"):
            with st.spinner("Integrando novas planilhas ao seu banco de dados principal..."):
                lista_dfs_novos = []
                for file in uploaded_files:
                    extensao = file.name.split('.')[-1].lower()
                    try:
                        if extensao == 'csv':
                            try:
                                df_bruto = pd.read_csv(file)
                            except:
                                file.seek(0)
                                df_bruto = pd.read_csv(file, sep=';', encoding='latin-1')
                        elif extensao == 'txt':
                            df_bruto = pd.read_csv(file, sep='\t')
                        else:
                            df_bruto = pd.read_excel(file)
                        
                        df_limpo = processar_planilha(df_bruto)
                        lista_dfs_novos.append(df_limpo)
                    except Exception as e:
                        st.error(f"Erro ao ler arquivo {file.name}.")

                if lista_dfs_novos:
                    df_novos_juntos = pd.concat(lista_dfs_novos, ignore_index=True)
                    
                    if os.path.exists(MASTER_DB_FILE):
                        df_antigo = pd.read_csv(MASTER_DB_FILE, dtype=str).fillna("")
                        df_unificado = pd.concat([df_antigo, df_novos_juntos], ignore_index=True)
                        tamanho_banco_antigo = len(df_antigo)
                    else:
                        df_unificado = df_novos_juntos
                        tamanho_banco_antigo = 0
                    
                    df_final = df_unificado.drop_duplicates(subset=['Telefone'], keep='last')
                    df_final = df_final.sort_values(by=['Região/Bairro', 'Nome'])
                    
                    tamanho_banco_novo = len(df_final)
                    novos_adicionados = tamanho_banco_novo - tamanho_banco_antigo
                    duplicadas_ignoradas = len(df_novos_juntos) - novos_adicionados
                    
                    df_final.to_csv(MASTER_DB_FILE, index=False)
                    st.session_state.df_final = df_final
                    
                    st.success(f"✨ Sucesso! {novos_adicionados} contatos novos entraram na base. Dados de agendamento foram lidos e atualizados.")

    if st.session_state.df_final is not None:
        df_resultado = st.session_state.df_final
        
        st.subheader("📊 Pré-visualização da Base Completa")
        st.dataframe(df_resultado, use_container_width=True)
        
        excel_pronto_abas = gerar_excel_final(df_resultado)
        excel_pronto_unico = gerar_excel_unico(df_resultado)
        
        st.markdown("<br>", unsafe_allow_html=True)
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            st.download_button(
                label="📥 Baixar Planilha Separada", 
                data=excel_pronto_abas, 
                file_name="Radar_Politico_Abas.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
        with col_btn2:
            st.download_button(
                label="📥 Baixar Dados Completos (Com Nº)", 
                data=excel_pronto_unico, 
                file_name="Radar_Politico_Completo.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.markdown("---")
    with st.expander("⚠️ Opções Avançadas (Zerar Sistema)"):
        st.warning("Cuidado: Esta ação apagará permanentemente todos os contatos salvos no sistema.")
        if st.button("🗑️ Apagar Banco de Dados Geral"):
            if os.path.exists(MASTER_DB_FILE):
                os.remove(MASTER_DB_FILE)
            st.session_state.df_final = None
            st.rerun()

# ==========================================
# MÓDULO 3: CONTROLE DE CASTRAÇÃO E ADOÇÃO
# ==========================================
elif menu_selecionado == "🐾 Controle de Castração":
    st.title("Controle de Castração e Adoção")
    st.write("Gerencie os pedidos de ajuda. Os dados são salvos automaticamente.")
    
    st.markdown("<br>", unsafe_allow_html=True)

    with st.form("form_castracao", clear_on_submit=True):
        st.subheader("Novo Cadastro")
        
        tipo_input = st.selectbox("Selecione a Categoria:", ["Ajuda / Castrar", "Adoção / Adotante"])
        
        col_form1, col_form2 = st.columns(2)
        with col_form1:
            nome_input = st.text_input("Nome", placeholder="Ex: Maria da Silva")
        with col_form2:
            telefone_input = st.text_input("Telefone", placeholder="(21) 9XXXX-XXXX")
            
        col_btn_add, col_btn_clear = st.columns(2)
        with col_btn_add:
            submitted = st.form_submit_button("➕ Incluir Registro")
        with col_btn_clear:
            st.form_submit_button("🧹 Limpar Campos")
            
        if submitted:
            if nome_input:
                tel_formatado = formatar_telefone(telefone_input)
                nova_linha = pd.DataFrame([{"Tipo": tipo_input, "Nome": nome_input, "Telefone": tel_formatado}])
                st.session_state.df_castracao = pd.concat([st.session_state.df_castracao, nova_linha], ignore_index=True)
                st.session_state.df_castracao.to_csv(CASTRACAO_DB_FILE, index=False)
                st.success("Registro adicionado e salvo com sucesso!")
            else:
                st.warning("Preencha o Nome para incluir o registro.")

    st.markdown("---")
    st.subheader("Base de Registros")
    st.info("💡 **Dica:** Para **Alterar**, clique duas vezes na célula da tabela abaixo. Para **Excluir**, selecione a linha no quadrado à esquerda e aperte a lixeira!")

    df_atualizado = st.data_editor(
        st.session_state.df_castracao,
        num_rows="dynamic",
        use_container_width=True,
        key="editor_castracao"
    )
    
    if not df_atualizado.equals(st.session_state.df_castracao):
        st.session_state.df_castracao = df_atualizado
        st.session_state.df_castracao.to_csv(CASTRACAO_DB_FILE, index=False)

    if not st.session_state.df_castracao.empty:
        
        output_castracao = io.BytesIO()
        with pd.ExcelWriter(output_castracao, engine='openpyxl') as writer:
            header_font = Font(bold=True)
            
            # Aba 1: Ajuda / Castrar
            df_ajuda = st.session_state.df_castracao[st.session_state.df_castracao["Tipo"] == "Ajuda / Castrar"][["Nome", "Telefone"]]
            df_ajuda.to_excel(writer, index=False, startrow=1, sheet_name="Planilha1", header=False)
            ws1 = writer.sheets["Planilha1"]
            ws1.cell(row=1, column=1, value="ajuda/ castrar").font = header_font
            ws1.cell(row=2, column=1, value="Nome").font = header_font
            ws1.cell(row=2, column=2, value="Telefone").font = header_font
            ws1.column_dimensions['A'].width = 30
            ws1.column_dimensions['B'].width = 25
            
            # Aba 2: Adoção / Adotante
            df_adocao = st.session_state.df_castracao[st.session_state.df_castracao["Tipo"] == "Adoção / Adotante"][["Nome", "Telefone"]]
            df_adocao.to_excel(writer, index=False, startrow=1, sheet_name="Planilha2", header=False)
            ws2 = writer.sheets["Planilha2"]
            ws2.cell(row=1, column=1, value="adoção/adotante").font = header_font
            ws2.cell(row=2, column=1, value="Nome").font = header_font
            ws2.cell(row=2, column=2, value="Telefone").font = header_font
            ws2.column_dimensions['A'].width = 30
            ws2.column_dimensions['B'].width = 25
        
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            label="📥 Baixar Planilha de Controle", 
            data=output_castracao.getvalue(), 
            file_name="Controle_Castracao.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )